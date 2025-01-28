from selenium import webdriver                                   # for interaction with browser
from selenium.webdriver.common.by import By                      # for usind By command for search element
from selenium.webdriver.common.keys import Keys                  # for sending keys from keyboard
from selenium.webdriver.chrome.options import Options            # for using chrome options
from openpyxl import load_workbook                               # for using excel file for read / write
import time                                                      # for using delays
import pandas as pd                                              # for accessing xlsx files
import logging                                                   # for reading errror logs


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Load credentials from Excel

def load_credentials(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active  # Load the first sheet
    credentials = [
        (row[0], row[1]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] and row[1]
    ]
    return credentials

# Main script
def main():
    # Path to your Excel 
    emails_file_path = 'receptions.xlsx'
    credentials_file = 'credentials.xlsx'
    driver_path = "./chromedriver"

    # Load data from xlsx and transform its to list
    logging.info("Loading email addresses and credentials...")
    df = pd.read_excel(emails_file_path)
    email_addresses = df['Email'].tolist()
    all_emails = ", ".join(email_addresses)
    all_credentials = load_credentials(credentials_file)

    # Email details
    email_subject = input("Enter the email subject: ")
    email_body = input("Enter the email body: ")

    # Chrome options
    chrome_options = Options()
    chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")

    # Process each credential
    for email, password in all_credentials:
        logging.info(f"Processing login for {email}...")
        driver = webdriver.Chrome(executable_path=driver_path, options=chrome_options)

        try:
            # Navigate to Gmail login page
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(5)
            driver.get("https://accounts.google.com/v3/signin/identifier?ddm=1&flowEntry=ServiceLogin&flowName=GlifWebSignIn&hl=en-gb&ifkv=AVdkyDn7HwHwp8nsRVVl46KFKkXG-D93AWdrGAQ5mp6l6MUTEJYpXqt1KzEaGJEfAiy3C0WX6qwFAw&continue=https%3A%2F%2Fmail.google.com%2Fmail%2Fu%2F0%2F%23inbox")

            # Enter email
            email_field = driver.find_element(By.XPATH, '//input[@id="identifierId"]')
            email_field.clear()
            email_field.send_keys(email)
            time.sleep(5)
            email_field.send_keys(Keys.RETURN)
            time.sleep(10)

            # Enter password
            password_field = driver.find_element(By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')
            password_field.clear()
            password_field.send_keys(password)
            time.sleep(5)
            password_field.send_keys(Keys.RETURN)
            time.sleep(10)

            logging.info("Login successful.")

            # Open the first email
            driver.find_element(By.XPATH, '/html/body/div[6]/div[3]/div/div[2]/div[2]/div/div/div/div[2]/div/div[1]/div/div/div[8]/div/div[1]/div[2]/div/table/tbody/tr[1]').click()
            time.sleep(10)

            # Open spam folder
            open_spam_folder = driver.find_element(By.NAME, "q")
            open_spam_folder.clear()
            open_spam_folder.send_keys("in:spam")
            time.sleep(10)
            open_spam_folder.send_keys(Keys.RETURN)
            time.sleep(10)

            # Open first email in spam
            driver.find_element(By.XPATH, '/html/body/div[6]/div[3]/div/div[2]/div[2]/div/div/div/div[2]/div/div[1]/div/div[2]/div[4]/div[1]/div/table/tbody/tr[1]').click()
            time.sleep(10)

            # Report not spam
            driver.find_element(By.XPATH, '//*[@id=":4"]/div[3]/div[1]/div/div[3]/div/div').click()
            time.sleep(10)

            # Open compose email
            driver.find_element(By.XPATH, '/html/body/div[6]/div[3]/div/div[2]/div[1]/div[1]/div/div').click()
            time.sleep(10)

            # Add recipients
            driver.switch_to.active_element.send_keys(all_emails) # Load from List i.e. all_email and it will get it from receptions.xlsx file
            time.sleep(5)

            # Add subject
            subject_field = driver.find_element(By.NAME, "subjectbox")
            subject_field.clear()
            subject_field.send_keys(email_subject)
            time.sleep(5)
            subject_field.send_keys(Keys.TAB)
            time.sleep(5)

            # Add email body 
            driver.switch_to.active_element.send_keys(email_body)
            time.sleep(10)

            # Send email
            driver.switch_to.active_element.send_keys(Keys.CONTROL, Keys.ENTER)
            time.sleep(5)

            logging.info("Email sent successfully.")

            # Logout the gmail
            gmail_logout = driver.find_element(By.XPATH,'//*[@id="gb"]/div[2]/div[3]/div[1]/div[2]/div/a').click()
            time.sleep(8)

            # Logout
            driver.get("https://accounts.google.com/Logout?ec=GAdAwAE&hl=en")
            time.sleep(10)
            logging.info(f"Completed task for {email}.")

        except Exception as e:
            logging.error(f"An error occurred for {email}: {e}")

        finally:
            # Ensure browser is closed properly
            driver.quit()

if __name__ == "__main__":
    main()
